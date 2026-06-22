from flask import Flask, render_template, request, jsonify, redirect, url_for
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from database import db, User, Bond, AuditLog, Reconciliation, ReconciliationItem
from email_service import send_bond_notification, send_invite_email
from datetime import datetime, timedelta
import os
import json
import secrets
import threading
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-before-deploying')

db_url = os.environ.get('DATABASE_URL', 'sqlite:///bonds.db')
# Railway PostgreSQL URLs start with postgres://, SQLAlchemy needs postgresql://
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access Bond Tracker.'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def log_action(bond, action, old_data=None):
    entry = AuditLog(
        bond_id     = bond.id,
        bond_number = bond.bond_number,
        action      = action,
        changed_by  = current_user.username,
        old_values  = json.dumps(old_data) if old_data else None,
        new_values  = json.dumps(bond.to_dict()),
    )
    db.session.add(entry)


# ── Auth ───────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        data     = request.get_json() or request.form
        username = (data.get('username') or '').strip()
        password = data.get('password') or ''

        user = User.query.filter_by(username=username).first()
        if user and user.active and user.check_password(password):
            login_user(user, remember=True)
            if request.is_json:
                return jsonify({'ok': True})
            return redirect(url_for('index'))

        error = 'Invalid username or password.'
        if request.is_json:
            return jsonify({'ok': False, 'error': error}), 401
        return render_template('login.html', error=error)

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── Pages ──────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=current_user)


@app.route('/admin')
@login_required
def admin():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    users   = User.query.order_by(User.created_at).all()
    app_url = os.environ.get('APP_URL', request.host_url.rstrip('/'))
    return render_template('admin.html', user=current_user, users=users, app_url=app_url)


# ── Bond API ───────────────────────────────────────────────────────────

@app.route('/api/bonds', methods=['GET'])
@login_required
def get_bonds():
    bonds = Bond.query.order_by(Bond.created_at.desc()).all()
    return jsonify([b.to_dict() for b in bonds])


@app.route('/api/bonds', methods=['POST'])
@login_required
def create_bond():
    data = request.get_json()

    bond_num = data.get('bond_number', '').strip()
    if bond_num and Bond.query.filter_by(bond_number=bond_num).first():
        return jsonify({'error': 'Bond number already exists.'}), 400

    bond = Bond(
        bond_number          = data['bond_number'].strip(),
        bond_type            = data['bond_type'],
        principal            = data['principal'].strip(),
        obligee              = data['obligee'].strip(),
        project              = data.get('project', '').strip(),
        project_description  = data.get('project_description', '').strip(),
        surety               = data['surety'].strip(),
        bond_amount          = data.get('bond_amount') or None,
        bid_bond_percent     = data.get('bid_bond_percent') or None,
        bid_date             = data.get('bid_date') or None,
        expiration_date      = data.get('expiration_date') or None,
        decision_date        = data.get('decision_date') or None,
        status               = data.get('status', 'Pending'),
        notes                = data.get('notes', '').strip(),
        work_on_hand         = data.get('work_on_hand', '').strip(),
        work_on_hand_low     = bool(data.get('work_on_hand_low', False)),
        low_bid              = bool(data.get('low_bid', False)),
        created_by           = current_user.username,
        created_at           = datetime.utcnow(),
    )
    db.session.add(bond)
    db.session.flush()
    log_action(bond, 'created')
    db.session.commit()

    send_bond_notification('created', bond, current_user.username)
    return jsonify(bond.to_dict()), 201


@app.route('/api/bonds/<int:bond_id>', methods=['PUT'])
@login_required
def update_bond(bond_id):
    bond     = Bond.query.get_or_404(bond_id)
    data     = request.get_json()
    old_data = bond.to_dict()
    old_status = bond.status

    new_num = data.get('bond_number', bond.bond_number).strip()
    if new_num != bond.bond_number and Bond.query.filter_by(bond_number=new_num).first():
        return jsonify({'error': 'Bond number already exists.'}), 400

    bond.bond_number         = new_num
    bond.bond_type           = data.get('bond_type',           bond.bond_type)
    bond.principal           = data.get('principal',           bond.principal).strip()
    bond.obligee             = data.get('obligee',             bond.obligee).strip()
    bond.project             = data.get('project',             bond.project or '').strip()
    bond.project_description = data.get('project_description', bond.project_description or '').strip()
    bond.surety              = data.get('surety',              bond.surety).strip()
    bond.bond_amount         = data.get('bond_amount')         or bond.bond_amount
    bond.bid_bond_percent    = data.get('bid_bond_percent')    or None
    bond.bid_date            = data.get('bid_date')            or bond.bid_date
    bond.expiration_date     = data.get('expiration_date')     or None
    bond.decision_date       = data.get('decision_date')       or bond.decision_date
    bond.status              = data.get('status',              bond.status)
    bond.notes               = data.get('notes',               bond.notes or '').strip()
    bond.work_on_hand        = data.get('work_on_hand',        bond.work_on_hand or '').strip()
    bond.work_on_hand_low    = bool(data.get('work_on_hand_low', bond.work_on_hand_low or False))
    bond.low_bid             = bool(data.get('low_bid', bond.low_bid or False))
    bond.updated_by          = current_user.username
    bond.updated_at          = datetime.utcnow()

    log_action(bond, 'updated', old_data)
    db.session.commit()

    if bond.status != old_status:
        send_bond_notification('status_changed', bond, current_user.username, old_status=old_status)
    else:
        send_bond_notification('updated', bond, current_user.username)

    return jsonify(bond.to_dict())


@app.route('/api/bonds/<int:bond_id>', methods=['DELETE'])
@login_required
def delete_bond(bond_id):
    bond     = Bond.query.get_or_404(bond_id)
    old_data = bond.to_dict()

    entry = AuditLog(
        bond_id     = bond.id,
        bond_number = bond.bond_number,
        action      = 'deleted',
        changed_by  = current_user.username,
        old_values  = json.dumps(old_data),
    )
    db.session.add(entry)
    send_bond_notification('deleted', bond, current_user.username)
    db.session.delete(bond)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/bonds/<int:bond_id>/create-final', methods=['POST'])
@login_required
def create_final_bond(bond_id):
    orig = Bond.query.get_or_404(bond_id)
    bond = Bond(
        bond_number         = None,
        bond_type           = 'Final Bond',
        principal           = orig.principal,
        obligee             = orig.obligee,
        project             = orig.project,
        project_description = orig.project_description,
        surety              = orig.surety,
        bond_amount         = orig.bond_amount,
        bid_date            = orig.bid_date,
        decision_date       = None,
        status              = 'Pending',
        notes               = '',
        work_on_hand        = orig.work_on_hand,
        work_on_hand_low    = orig.work_on_hand_low,
        low_bid             = orig.low_bid,
        created_by          = current_user.username,
        created_at          = datetime.utcnow(),
    )
    db.session.add(bond)
    db.session.flush()
    log_action(bond, 'created')
    db.session.commit()
    send_bond_notification('created', bond, current_user.username)
    return jsonify(bond.to_dict()), 201


# ── Audit log API ──────────────────────────────────────────────────────

@app.route('/api/audit')
@login_required
def get_audit():
    logs = AuditLog.query.order_by(AuditLog.changed_at.desc()).limit(500).all()
    return jsonify([l.to_dict() for l in logs])


# ── Reconciliation API ─────────────────────────────────────────────────

@app.route('/api/reconciliations', methods=['GET'])
@login_required
def get_reconciliations():
    recs = Reconciliation.query.order_by(Reconciliation.created_at.desc()).all()
    return jsonify([r.to_dict() for r in recs])


@app.route('/api/reconciliations', methods=['POST'])
@login_required
def create_reconciliation():
    data = request.get_json()
    rec = Reconciliation(
        carrier    = data.get('carrier', '').strip(),
        period     = data.get('period', ''),
        status     = data.get('status', 'In Progress'),
        notes      = data.get('notes', '').strip(),
        created_by = current_user.username,
    )
    db.session.add(rec)
    db.session.commit()
    return jsonify(rec.to_dict(include_items=True)), 201


@app.route('/api/reconciliations/<int:rec_id>', methods=['GET'])
@login_required
def get_reconciliation(rec_id):
    rec = Reconciliation.query.get_or_404(rec_id)
    return jsonify(rec.to_dict(include_items=True))


@app.route('/api/reconciliations/<int:rec_id>', methods=['PUT'])
@login_required
def update_reconciliation(rec_id):
    rec = Reconciliation.query.get_or_404(rec_id)
    data = request.get_json()
    rec.carrier = data.get('carrier', rec.carrier).strip()
    rec.period  = data.get('period',  rec.period)
    rec.status  = data.get('status',  rec.status)
    rec.notes   = data.get('notes',   rec.notes or '').strip()
    db.session.commit()
    return jsonify(rec.to_dict(include_items=True))


@app.route('/api/reconciliations/<int:rec_id>', methods=['DELETE'])
@login_required
def delete_reconciliation(rec_id):
    rec = Reconciliation.query.get_or_404(rec_id)
    db.session.delete(rec)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/reconciliations/<int:rec_id>/items', methods=['POST'])
@login_required
def add_recon_item(rec_id):
    Reconciliation.query.get_or_404(rec_id)
    data = request.get_json()
    ca = data.get('carrier_amount') or None
    oa = data.get('our_amount') or None
    if ca is not None: ca = float(ca)
    if oa is not None: oa = float(oa)
    auto_status = 'Pending'
    if ca is not None and oa is not None:
        auto_status = 'Matched' if abs(ca - oa) < 0.01 else 'Discrepancy'
    elif data.get('status') == 'Missing':
        auto_status = 'Missing'
    item = ReconciliationItem(
        recon_id       = rec_id,
        bond_number    = data.get('bond_number', '').strip(),
        principal      = data.get('principal', '').strip(),
        carrier_amount = ca,
        our_amount     = oa,
        status         = data.get('status') or auto_status,
        notes          = data.get('notes', '').strip(),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/reconciliations/<int:rec_id>/items/<int:item_id>', methods=['PUT'])
@login_required
def update_recon_item(rec_id, item_id):
    item = ReconciliationItem.query.filter_by(id=item_id, recon_id=rec_id).first_or_404()
    data = request.get_json()
    ca = data.get('carrier_amount') or None
    oa = data.get('our_amount') or None
    if ca is not None: ca = float(ca)
    if oa is not None: oa = float(oa)
    item.bond_number    = data.get('bond_number', item.bond_number or '').strip()
    item.principal      = data.get('principal',   item.principal or '').strip()
    item.carrier_amount = ca if ca is not None else item.carrier_amount
    item.our_amount     = oa if oa is not None else item.our_amount
    if ca is not None and oa is not None:
        item.status = 'Matched' if abs(ca - oa) < 0.01 else 'Discrepancy'
    item.status = data.get('status') or item.status
    item.notes  = data.get('notes', item.notes or '').strip()
    db.session.commit()
    return jsonify(item.to_dict())


@app.route('/api/reconciliations/<int:rec_id>/items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_recon_item(rec_id, item_id):
    item = ReconciliationItem.query.filter_by(id=item_id, recon_id=rec_id).first_or_404()
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


# ── User API (admin only) ──────────────────────────────────────────────

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()

    if User.query.filter_by(username=data.get('username', '').strip()).first():
        return jsonify({'error': 'Username already exists.'}), 400
    if User.query.filter_by(email=data.get('email', '').strip()).first():
        return jsonify({'error': 'Email already exists.'}), 400

    user = User(
        username = data['username'].strip(),
        email    = data['email'].strip(),
        role     = data.get('role', 'user'),
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    user = User.query.get_or_404(user_id)
    data = request.get_json()

    user.email  = data.get('email', user.email).strip()
    user.role   = data.get('role', user.role)
    user.active = data.get('active', user.active)
    if data.get('password'):
        user.set_password(data['password'])
    db.session.commit()
    return jsonify(user.to_dict())


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    if user_id == current_user.id:
        return jsonify({'error': 'You cannot delete your own account.'}), 400
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/users/invite', methods=['POST'])
@login_required
def invite_user():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()

    if User.query.filter_by(username=data.get('username', '').strip()).first():
        return jsonify({'error': 'Username already exists.'}), 400
    if User.query.filter_by(email=data.get('email', '').strip()).first():
        return jsonify({'error': 'Email already exists.'}), 400

    token = secrets.token_urlsafe(32)
    user  = User(
        username       = data['username'].strip(),
        email          = data['email'].strip(),
        role           = data.get('role', 'user'),
        active         = False,
        invite_token   = token,
        invite_expires = datetime.utcnow() + timedelta(hours=72),
    )
    db.session.add(user)
    db.session.commit()

    invite_link = url_for('accept_invite', token=token, _external=True)
    threading.Thread(target=send_invite_email, args=(user.email, user.username, invite_link), daemon=True).start()

    return jsonify({'ok': True, 'invite_link': invite_link}), 201


@app.route('/invite/<token>', methods=['GET', 'POST'])
def accept_invite(token):
    user = User.query.filter_by(invite_token=token).first()

    if not user:
        return render_template('accept_invite.html', error='This invite link is invalid.')
    if user.invite_expires and datetime.utcnow() > user.invite_expires:
        return render_template('accept_invite.html', error='This invite link has expired. Ask your administrator to resend the invite.')
    if user.password_hash:
        return render_template('accept_invite.html', error='This invite has already been used. Please log in.')

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm', '')
        if len(password) < 8:
            return render_template('accept_invite.html', token=token, username=user.username,
                                   error='Password must be at least 8 characters.')
        if password != confirm:
            return render_template('accept_invite.html', token=token, username=user.username,
                                   error='Passwords do not match.')

        user.set_password(password)
        user.active         = True
        user.invite_token   = None
        user.invite_expires = None
        db.session.commit()
        login_user(user)
        return redirect(url_for('index'))

    return render_template('accept_invite.html', token=token, username=user.username)


@app.route('/api/users/<int:user_id>/resend-invite', methods=['POST'])
@login_required
def resend_invite(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    user = User.query.get_or_404(user_id)
    if user.password_hash:
        return jsonify({'error': 'User has already accepted their invite.'}), 400

    token = secrets.token_urlsafe(32)
    user.invite_token   = token
    user.invite_expires = datetime.utcnow() + timedelta(hours=72)
    db.session.commit()

    invite_link = url_for('accept_invite', token=token, _external=True)
    threading.Thread(target=send_invite_email, args=(user.email, user.username, invite_link), daemon=True).start()

    return jsonify({'ok': True})


@app.route('/reports/pdf')
@login_required
def generate_report():
    from fpdf import FPDF
    from flask import make_response

    status_filter = request.args.get('status', '')
    type_filter   = request.args.get('bond_type', '')
    from_date     = request.args.get('from_date', '')
    to_date       = request.args.get('to_date', '')

    query = Bond.query
    if status_filter:
        query = query.filter(Bond.status == status_filter)
    if type_filter:
        query = query.filter(Bond.bond_type == type_filter)
    if from_date:
        query = query.filter(Bond.bid_date >= from_date)
    if to_date:
        query = query.filter(Bond.bid_date <= to_date)

    bonds = query.order_by(Bond.bid_date.asc(), Bond.bond_number.asc()).all()

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    # ── Header ──────────────────────────────────────────────
    pdf.set_font('Helvetica', 'B', 18)
    pdf.set_text_color(13, 27, 42)
    pdf.cell(0, 10, 'IAMA BondDesk', align='C')
    pdf.ln(10)

    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(95, 99, 104)
    pdf.cell(0, 6, 'Insurance Agency of Mid America', align='C')
    pdf.ln(6)

    pdf.ln(2)

    filter_parts = [status_filter if status_filter else 'All Statuses']
    if from_date or to_date:
        d = ''
        if from_date: d += f'From {from_date}'
        if to_date:   d += f'  To {to_date}'
        filter_parts.append(d.strip())

    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(13, 27, 42)
    pdf.cell(0, 8, f'Bond Report  -  {", ".join(filter_parts)}', align='C')
    pdf.ln(8)

    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(95, 99, 104)
    pdf.cell(0, 5, f'Generated {datetime.utcnow().strftime("%m/%d/%Y %I:%M %p")} UTC  |  {len(bonds)} bond(s)', align='C')
    pdf.ln(8)

    # ── Table header ────────────────────────────────────────
    cols = [
        ('Bond #',   24), ('Type',     24), ('Principal', 42),
        ('Obligee',  34), ('Project',  40), ('Surety',    30),
        ('Amount',   22), ('Bid Date', 22), ('Status',    29),
    ]

    pdf.set_fill_color(13, 27, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    for label, w in cols:
        pdf.cell(w, 8, label, fill=True, align='C')
    pdf.ln(8)

    # ── Rows ────────────────────────────────────────────────
    status_colors = {
        'Approved':     (46, 125, 50),
        'Not Approved': (198, 40, 40),
        'Pending':      (230, 81, 0),
        'Final Bond':   (13, 71, 161),
    }
    pdf.set_font('Helvetica', '', 8)

    for i, bond in enumerate(bonds):
        fill = i % 2 == 0
        if fill:
            pdf.set_fill_color(248, 249, 250)

        amt = f'${bond.bond_amount:,.0f}' if bond.bond_amount else '-'
        if bond.bond_amount and bond.bid_bond_percent and bond.bond_type == 'Bid Bond':
            pct_amt = bond.bond_amount * bond.bid_bond_percent / 100
            amt += f' ({bond.bid_bond_percent:.0f}%=${pct_amt:,.0f})'
        bid = bond.bid_date or '-'

        def safe(s):
            return ''.join(c if ord(c) < 256 else '?' for c in (s or ''))

        def trunc(s, n):
            s = safe(s)
            return s[:n-1] + '...' if len(s) > n else s

        row_vals = [
            (trunc(bond.bond_number or 'N/A', 12), cols[0][1], 'L'),
            (trunc(bond.bond_type,   12),           cols[1][1], 'L'),
            (trunc(bond.principal,   23),           cols[2][1], 'L'),
            (trunc(bond.obligee,     19),           cols[3][1], 'L'),
            (trunc(bond.project or '', 22),         cols[4][1], 'L'),
            (trunc(bond.surety,      16),           cols[5][1], 'L'),
            (trunc(amt,              12),           cols[6][1], 'R'),
            (bid,                                   cols[7][1], 'C'),
        ]

        pdf.set_text_color(60, 64, 67)
        for text, w, align in row_vals:
            pdf.cell(w, 7, text, border='B', fill=fill, align=align)

        sc = status_colors.get(bond.status, (60, 64, 67))
        pdf.set_text_color(*sc)
        pdf.cell(cols[8][1], 7, bond.status, border='B', fill=fill, align='C')
        pdf.set_text_color(60, 64, 67)
        pdf.ln(7)

    # ── Summary ─────────────────────────────────────────────
    pdf.ln(4)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(95, 99, 104)
    approved = sum(1 for b in bonds if b.status == 'Approved')
    not_appr = sum(1 for b in bonds if b.status == 'Not Approved')
    pending  = sum(1 for b in bonds if b.status == 'Pending')
    woh_low  = sum(1 for b in bonds if b.work_on_hand_low)
    pdf.cell(0, 6,
        f'Total: {len(bonds)}   |   Approved: {approved}   |   Not Approved: {not_appr}   |   Pending: {pending}   |   Work on Hand Low: {woh_low}',
        align='C')

    pdf_bytes = bytes(pdf.output())
    fname = f'IAMA-BondReport-{datetime.utcnow().strftime("%Y%m%d")}.pdf'
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


@app.route('/reports/excel')
@login_required
def generate_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import make_response
    import io

    status_filter = request.args.get('status', '')
    type_filter   = request.args.get('bond_type', '')
    from_date     = request.args.get('from_date', '')
    to_date       = request.args.get('to_date', '')

    query = Bond.query
    if status_filter:
        query = query.filter(Bond.status == status_filter)
    if type_filter:
        query = query.filter(Bond.bond_type == type_filter)
    if from_date:
        query = query.filter(Bond.bid_date >= from_date)
    if to_date:
        query = query.filter(Bond.bid_date <= to_date)
    bonds = query.order_by(Bond.bid_date.asc(), Bond.bond_number.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bond Report'

    headers = ['Bond #', 'Bond Type', 'Principal', 'Obligee', 'Project', 'Project Description',
               'Surety', 'Amount', 'Bid Bond %', 'Bid Date', 'Decision Date',
               'Status', 'Notes', 'Work on Hand', 'Work on Hand Low']

    hdr_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    status_fills = {
        'Approved':     PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'Not Approved': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
        'Pending':      PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    }

    for ri, bond in enumerate(bonds, 2):
        row_data = [
            bond.bond_number or 'N/A', bond.bond_type, bond.principal,
            bond.obligee, bond.project or '', bond.project_description or '',
            bond.surety,
            bond.bond_amount,
            f'{bond.bid_bond_percent}%' if bond.bid_bond_percent else '',
            bond.bid_date, bond.decision_date,
            bond.status, bond.notes or '',
            bond.work_on_hand or '',
            'Yes' if bond.work_on_hand_low else 'No',
        ]
        fill = status_fills.get(bond.status)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'IAMA-BondReport-{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    return resp


@app.route('/api/bonds/extract', methods=['POST'])
@login_required
def extract_bond_pdf():
    import pdfplumber
    import anthropic as _anthropic

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded.'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'File must be a PDF.'}), 400

    try:
        with pdfplumber.open(file) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
    except Exception as e:
        return jsonify({'error': f'Could not read PDF: {e}'}), 400

    if not text.strip():
        return jsonify({'error': 'No text found in this PDF — it may be a scanned image.'}), 400

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY not set in Railway variables.'}), 500

    client = _anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-haiku-4-5-20251001',
        max_tokens=1024,
        messages=[{
            'role': 'user',
            'content': (
                'Extract bond information from this document. '
                'Return ONLY a JSON object with these exact keys:\n'
                '  bond_type (one of: Bid Bond / Final Bond / License Bond / Maintenance Bond / Other),\n'
                '  principal, obligee, surety,\n'
                '  bond_amount: the CONTRACT amount or bid amount (NOT the bond penalty or penal sum — the full contract/project dollar value, number only no $ or commas, or null),\n'
                '  bid_bond_percent: the bid bond percentage (e.g. 5, 10, or 20 — number only, or null if not stated),\n'
                '  bid_date (YYYY-MM-DD or null), decision_date (YYYY-MM-DD or null),\n'
                '  project: the name of the project or job this bond is for (short name or title, or null),\n'
                '  project_description: a brief description of the project scope or work (1-3 sentences, or null)\n'
                'Return only the JSON, no extra text.\n\n'
                f'Document:\n{text[:4000]}'
            )
        }]
    )

    try:
        raw = message.content[0].text.strip()
        print(f'[extract] Raw AI response: {raw[:500]}')
        # Strip markdown code fences if present
        if '```' in raw:
            raw = raw.split('```')[1]
            if raw.startswith('json'):
                raw = raw[4:]
        # Find JSON object within the text
        start = raw.find('{')
        end   = raw.rfind('}')
        if start != -1 and end != -1:
            raw = raw[start:end+1]
        extracted = json.loads(raw.strip())
    except Exception as e:
        print(f'[extract] Parse error: {e} | raw: {message.content[0].text[:300]}')
        return jsonify({'error': 'Could not parse AI response. Try again.'}), 500

    return jsonify(extracted)


@app.route('/api/me/password', methods=['PUT'])
@login_required
def change_password():
    data = request.get_json()
    if not current_user.check_password(data.get('current_password', '')):
        return jsonify({'error': 'Current password is incorrect.'}), 400
    current_user.set_password(data['new_password'])
    db.session.commit()
    return jsonify({'ok': True})


# ── Startup ────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    # Allow bond_number to be nullable (migration for existing databases)
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            if 'postgresql' in str(db.engine.url):
                conn.execute(text('ALTER TABLE bonds ALTER COLUMN bond_number DROP NOT NULL'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS expiration_date VARCHAR(10)'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS project VARCHAR(300)'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS project_description TEXT'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS bid_bond_percent FLOAT'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS work_on_hand TEXT'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS work_on_hand_low BOOLEAN DEFAULT FALSE'))
                conn.execute(text('ALTER TABLE bonds ADD COLUMN IF NOT EXISTS low_bid BOOLEAN DEFAULT FALSE'))
                conn.commit()
    except Exception:
        pass
    if not User.query.first():
        _admin = User(username='admin', email='admin@company.com', role='admin', active=True)
        _admin.set_password('changeme123')
        db.session.add(_admin)
        db.session.commit()
        print('Default admin created — username: admin  password: changeme123')
        print('IMPORTANT: Change this password immediately after first login.')


if __name__ == '__main__':
    app.run(debug=True, port=5000)
