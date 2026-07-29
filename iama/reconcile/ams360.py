import re

import pdfplumber

TAIL_RE = re.compile(
    r'(-?[\d,]+\.\d{2})\s+([\d.]+)\s*%\s+([A-Za-z])\s+'
    r'(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+(YES|NO)\s*$'
)
DATE_RE = re.compile(r'(\d{1,2}/\d{1,2}/\d{2,4})')
LOB_PREFIXES = ["BUA", "CUE", "PMT", "WC", "CNP", "IM", "C"]


def _to_num(s):
    s = str(s).strip()
    neg = s.startswith('(') and s.endswith(')')
    s = s.replace('(', '').replace(')', '').replace(',', '').replace('$', '')
    try:
        n = float(s)
    except ValueError:
        return 0.0
    return -n if neg else n


def _cluster_lines(page, tolerance=3):
    # use_text_flow=True: this AMS360 export has a font whose per-character advance
    # widths pdfminer computes incorrectly for a couple of adjacent columns (client
    # name suffix / policy number end up at nearly identical x0 with default word
    # extraction, scrambling their characters together). text-flow order avoids that
    # and, once words are properly segmented, sorting by x0 gives the correct order
    # for every other column (verified against this real file).
    words = page.extract_words(use_text_flow=True)
    items = [{'x': w['x0'], 'y': w['top'], 'text': w['text']} for w in words]
    items.sort(key=lambda w: (w['y'], w['x']))
    rows = []
    current = None
    for it in items:
        if current is not None and abs(current['anchorY'] - it['y']) <= tolerance:
            current['items'].append(it)
        else:
            current = {'anchorY': it['y'], 'items': [it]}
            rows.append(current)
    lines = []
    for r in rows:
        r['items'].sort(key=lambda w: w['x'])
        text = ' '.join(w['text'] for w in r['items'])
        text = re.sub(r'\s+', ' ', text).strip()
        if text:
            lines.append(text)
    return lines


def extract_pdf_lines(file_obj):
    """Returns (lines, total_words) for a PDF file-like object."""
    lines = []
    total_words = 0
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            total_words += len(page.extract_words())
            lines.extend(_cluster_lines(page))
    return lines, total_words


def parse_lines(lines):
    """Parse AMS360 Direct Bill Statement text lines into transaction rows."""
    rows = []
    for text in lines:
        tail_match = TAIL_RE.search(text)
        if not tail_match:
            continue
        prem_fee, agcy_pct, _amd, agcy_comm, _comm_bal, _je_adjust, _comm_rcvd, _rec = tail_match.groups()
        head = text[:tail_match.start()].strip()

        date_match = DATE_RE.search(head)
        if not date_match:
            continue
        before_date = head[:date_match.start()].strip()

        policy_match = re.search(r'(\d{6,})', before_date)
        if not policy_match:
            continue
        policy_core = policy_match.group(1)
        name_part = before_date[:policy_match.start()].strip()
        for p in LOB_PREFIXES:
            if name_part.endswith(' ' + p):
                name_part = name_part[:-len(p)].strip()
                break
            if name_part.endswith(p) and len(p) > 1:
                name_part = name_part[:-len(p)].strip()
                break

        rows.append({
            'name': name_part,
            'policy_core': policy_core,
            'prem_fee': _to_num(prem_fee),
            'agcy_pct': _to_num(agcy_pct),
            'agcy_comm': _to_num(agcy_comm),
        })
    return rows


def parse_pdf(file_obj):
    lines, total_words = extract_pdf_lines(file_obj)
    return parse_lines(lines)


def parse_spreadsheet(file_obj, filename):
    """Parse an AMS360 export given as .xlsx/.xls/.csv."""
    rows = []
    if filename.lower().endswith('.csv'):
        import csv
        import io
        text = file_obj.read()
        if isinstance(text, bytes):
            text = text.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        records = list(reader)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        records = []
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            records.append(dict(zip(header, values)))

    def find_key(keys, needle):
        for k in keys:
            if needle in re.sub(r'[^a-z]', '', k.lower()):
                return k
        return None

    for record in records:
        keys = list(record.keys())
        name_key = find_key(keys, 'name') or find_key(keys, 'insured')
        policy_key = find_key(keys, 'policy')
        comm_key = find_key(keys, 'agcycomm') or find_key(keys, 'commission')
        prem_key = find_key(keys, 'premfee') or find_key(keys, 'premium')
        if not name_key or not policy_key or not comm_key:
            continue
        policy_match = re.search(r'(\d{6,})', str(record[policy_key]))
        if not policy_match:
            continue
        rows.append({
            'name': str(record[name_key]).strip(),
            'policy_core': policy_match.group(1),
            'prem_fee': _to_num(record.get(prem_key, 0) or 0),
            'agcy_comm': _to_num(record[comm_key]),
        })
    return rows
